#!/usr/bin/env python3
"""
Script để xóa hết users cũ và tạo lại từ file Excel mới
"""
import openpyxl
import requests
from datetime import datetime
import time

EXCEL_FILE = '/home/user/uploaded_files/DANH SACH TAI KHOAN.xlsx'
API_BASE = 'http://localhost:3000'

# Mapping vị trí -> position_id
POSITION_MAPPING = {
    'PTGĐ': 1,
    'GĐKD': 2,
    'GĐKDCC': 2,  # Cũng là GĐKD
    'Trợ lý Kinh Doanh': 3,
    'Giám sát': 4,
    'Giám Sát': 4  # Fix typo
}

# Mapping khối -> region_id
REGION_MAPPING = {
    'Bình Dương': 1,
    'Hà Nội': 2,
    'Miền Trung': 3,
    'Hồ Chí Minh': 4
}

def read_excel_data():
    """Đọc dữ liệu từ Excel"""
    print("📖 Đọc file Excel...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    data = []
    headers = [cell.value for cell in ws[1]]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                if isinstance(value, datetime):
                    row_dict[headers[i]] = value.strftime('%Y-%m-%d')
                else:
                    row_dict[headers[i]] = value
        
        if row_dict.get('Tên đăng nhập'):
            data.append(row_dict)
    
    print(f"✅ Đọc được {len(data)} users từ Excel")
    return data

def delete_all_users():
    """Xóa tất cả users (trừ admin)"""
    print("\n🗑️  Xóa hết users cũ...")
    
    # Get list of all users
    response = requests.get(f'{API_BASE}/api/admin/users')
    users = response.json()['users']
    
    deleted_count = 0
    for user in users:
        try:
            requests.delete(f'{API_BASE}/api/admin/users/{user["id"]}')
            deleted_count += 1
            if deleted_count % 10 == 0:
                print(f"  Đã xóa {deleted_count}/{len(users)} users...")
        except Exception as e:
            print(f"  ⚠️  Không thể xóa user {user['id']}: {e}")
    
    print(f"✅ Đã xóa {deleted_count} users")

def create_admin_accounts():
    """Tạo 4 tài khoản admin"""
    print("\n👑 Tạo 4 tài khoản admin...")
    
    # Admin accounts
    admins = [
        ('admin', 'admin123', 'Administrator'),
        ('admin1', 'admin123', 'Administrator 1'),
        ('admin2', 'admin123', 'Administrator 2'),
        ('admin3', 'admin123', 'Administrator 3')
    ]
    
    # First, delete existing admin users via database
    import subprocess
    
    # Delete all existing users except we'll recreate admin
    subprocess.run([
        'npx', 'wrangler', 'd1', 'execute', 'webapp-production', '--local',
        '--command', 'DELETE FROM users WHERE id > 0;'
    ], cwd='/home/user/webapp', capture_output=True)
    
    print("  Đã xóa tất cả users trong database")
    
    # Create 4 admin accounts
    for username, password, full_name in admins:
        try:
            subprocess.run([
                'npx', 'wrangler', 'd1', 'execute', 'webapp-production', '--local',
                '--command', 
                f"INSERT INTO users (username, password, full_name, region_id, position_id, start_date) "
                f"VALUES ('{username}', '{password}', '{full_name}', 1, 1, '2024-01-01');"
            ], cwd='/home/user/webapp', capture_output=True, check=True)
            
            print(f"  ✅ Tạo {username} thành công")
        except Exception as e:
            print(f"  ⚠️  Lỗi tạo {username}: {e}")
    
    print(f"✅ Đã tạo {len(admins)} admin accounts")

def create_users_from_excel(data):
    """Tạo users từ dữ liệu Excel"""
    print(f"\n👥 Tạo {len(data)} users từ Excel...")
    
    import subprocess
    
    success_count = 0
    error_count = 0
    
    for idx, record in enumerate(data, 1):
        username = record.get('Tên đăng nhập', '').strip()
        full_name = record.get('Họ và tên', '').strip()
        khoi = record.get('Khối', '').strip()
        vi_tri = record.get('Vị Trí', '').strip()
        start_date = record.get('Ngày nhận việc', '2024-01-01')
        
        # Skip if missing essential fields
        if not username or not full_name:
            print(f"  ⚠️  [{idx}/{len(data)}] Bỏ qua: thiếu username hoặc họ tên")
            error_count += 1
            continue
        
        # Map to IDs
        region_id = REGION_MAPPING.get(khoi)
        position_id = POSITION_MAPPING.get(vi_tri)
        
        if not region_id:
            print(f"  ⚠️  [{idx}/{len(data)}] {full_name}: Không tìm thấy khối '{khoi}'")
            error_count += 1
            continue
        
        if not position_id:
            print(f"  ⚠️  [{idx}/{len(data)}] {full_name}: Không tìm thấy vị trí '{vi_tri}'")
            error_count += 1
            continue
        
        # Create user via database
        try:
            cmd = [
                'npx', 'wrangler', 'd1', 'execute', 'webapp-production', '--local',
                '--command',
                f"INSERT INTO users (username, password, full_name, region_id, position_id, start_date) "
                f"VALUES ('{username}', '123456', '{full_name}', {region_id}, {position_id}, '{start_date}');"
            ]
            
            result = subprocess.run(cmd, cwd='/home/user/webapp', capture_output=True, text=True)
            
            if result.returncode == 0:
                success_count += 1
                if success_count % 10 == 0:
                    print(f"  ✅ [{success_count}/{len(data)}] Đã tạo {success_count} users...")
            else:
                error_count += 1
                if 'UNIQUE constraint failed' in result.stderr:
                    print(f"  ⚠️  [{idx}/{len(data)}] {username}: Username đã tồn tại")
                else:
                    print(f"  ⚠️  [{idx}/{len(data)}] {username}: Lỗi tạo user")
        
        except Exception as e:
            error_count += 1
            print(f"  ⚠️  [{idx}/{len(data)}] {username}: {str(e)[:50]}")
    
    print(f"\n✅ Tạo thành công: {success_count}/{len(data)} users")
    print(f"❌ Lỗi: {error_count}/{len(data)} users")
    
    return success_count, error_count

def verify_users():
    """Kiểm tra số lượng users đã tạo"""
    print("\n📊 Kiểm tra kết quả...")
    
    import subprocess
    result = subprocess.run([
        'npx', 'wrangler', 'd1', 'execute', 'webapp-production', '--local',
        '--command',
        "SELECT COUNT(*) as total FROM users;"
    ], cwd='/home/user/webapp', capture_output=True, text=True)
    
    print(result.stdout)
    
    # Group by position
    result = subprocess.run([
        'npx', 'wrangler', 'd1', 'execute', 'webapp-production', '--local',
        '--command',
        """
        SELECT 
            p.display_name as position,
            r.name as region,
            COUNT(*) as count
        FROM users u
        JOIN positions p ON u.position_id = p.id
        JOIN regions r ON u.region_id = r.id
        WHERE u.username NOT LIKE 'admin%'
        GROUP BY p.id, r.id
        ORDER BY r.id, p.id
        """
    ], cwd='/home/user/webapp', capture_output=True, text=True)
    
    print("\n📈 Thống kê theo Khối và Vị trí:")
    print(result.stdout)

if __name__ == '__main__':
    print("=" * 80)
    print("🔄 TẠO LẠI USERS TỪ FILE EXCEL MỚI")
    print("=" * 80)
    
    # Read Excel data
    data = read_excel_data()
    
    # Create admin accounts (this will delete all users first)
    create_admin_accounts()
    
    # Create users from Excel
    success, error = create_users_from_excel(data)
    
    # Verify
    verify_users()
    
    print("\n" + "=" * 80)
    print("✅ HOÀN TẤT!")
    print("=" * 80)
    print(f"📊 Tổng kết:")
    print(f"   • Admin accounts: 4")
    print(f"   • Users thành công: {success}")
    print(f"   • Users lỗi: {error}")
    print(f"   • Tổng: {4 + success}")
    print(f"\n🔑 Admin accounts:")
    print(f"   • admin / admin123")
    print(f"   • admin1 / admin123")
    print(f"   • admin2 / admin123")
    print(f"   • admin3 / admin123")
    print(f"\n🔐 User password: 123456")
    print("=" * 80)
