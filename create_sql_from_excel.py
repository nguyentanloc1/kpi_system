#!/usr/bin/env python3
"""
Tạo SQL script để recreate users
"""
import openpyxl
from datetime import datetime

EXCEL_FILE = '/home/user/uploaded_files/DANH SACH TAI KHOAN.xlsx'

# Mapping
POSITION_MAPPING = {
    'PTGĐ': 1,
    'GĐKD': 2,
    'GĐKDCC': 2,
    'Trợ lý Kinh Doanh': 3,
    'Giám sát': 4,
    'Giám Sát': 4
}

REGION_MAPPING = {
    'Bình Dương': 1,
    'Hà Nội': 2,
    'Miền Trung': 3,
    'Hồ Chí Minh': 4
}

def read_excel_data():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    data = []
    headers = [cell.value for cell in ws[1]]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                if isinstance(value, datetime):
                    row_dict[headers[i]] = value.strftime('%Y-%m-%d')
                else:
                    row_dict[headers[i]] = value
        
        if row_dict.get('Tên đăng nhập'):
            data.append(row_dict)
    
    return data

# Read data
data = read_excel_data()

print(f"📖 Đọc được {len(data)} users từ Excel\n")

# Create SQL
sql = []

# 1. Delete all users
sql.append("-- Xóa tất cả users hiện tại")
sql.append("DELETE FROM users WHERE id > 0;")
sql.append("")

# 2. Create 4 admin accounts
sql.append("-- Tạo 4 admin accounts")
sql.append("INSERT INTO users (username, password, full_name, region_id, position_id, start_date, employee_id, team) VALUES")
sql.append("  ('admin', 'admin123', 'Administrator', 1, 1, '2024-01-01', 'ADMIN001', 'ADMIN'),")
sql.append("  ('admin1', 'admin123', 'Administrator 1', 1, 1, '2024-01-01', 'ADMIN002', 'ADMIN'),")
sql.append("  ('admin2', 'admin123', 'Administrator 2', 1, 1, '2024-01-01', 'ADMIN003', 'ADMIN'),")
sql.append("  ('admin3', 'admin123', 'Administrator 3', 1, 1, '2024-01-01', 'ADMIN004', 'ADMIN');")
sql.append("")

# 3. Create users from Excel
sql.append("-- Tạo users từ Excel")
sql.append("INSERT INTO users (username, password, full_name, region_id, position_id, start_date, employee_id, team) VALUES")

values = []
success = 0
errors = []

for idx, record in enumerate(data, 1):
    username = record.get('Tên đăng nhập', '').strip()
    full_name = record.get('Họ và tên', '').strip()
    khoi = record.get('Khối', '').strip()
    vi_tri = record.get('Vị Trí', '').strip()
    start_date = record.get('Ngày nhận việc', '2024-01-01')
    employee_id = str(record.get('MSNV', '')).strip()
    team = str(record.get('Team', '')).strip()
    
    if not username or not full_name:
        errors.append(f"Row {idx}: Missing username or name")
        continue
    
    region_id = REGION_MAPPING.get(khoi)
    position_id = POSITION_MAPPING.get(vi_tri)
    
    if not region_id:
        errors.append(f"Row {idx} ({full_name}): Unknown region '{khoi}'")
        continue
    
    if not position_id:
        errors.append(f"Row {idx} ({full_name}): Unknown position '{vi_tri}'")
        continue
    
    # Escape single quotes in names
    full_name_escaped = full_name.replace("'", "''")
    
    values.append(f"  ('{username}', '123456', '{full_name_escaped}', {region_id}, {position_id}, '{start_date}', '{employee_id}', '{team}')")
    success += 1

sql.append(",\n".join(values) + ";")

# Write SQL file
with open('/home/user/webapp/recreate_users.sql', 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql))

print(f"✅ Tạo SQL file thành công!")
print(f"   • File: /home/user/webapp/recreate_users.sql")
print(f"   • Admin accounts: 4")
print(f"   • Users: {success}/{len(data)}")

if errors:
    print(f"\n⚠️  Có {len(errors)} lỗi:")
    for err in errors[:10]:
        print(f"   • {err}")
    if len(errors) > 10:
        print(f"   ... và {len(errors) - 10} lỗi khác")

# Statistics
print(f"\n📊 Thống kê:")
stats = {}
for record in data:
    khoi = record.get('Khối', 'N/A')
    vi_tri = record.get('Vị Trí', 'N/A')
    
    region_id = REGION_MAPPING.get(khoi)
    position_id = POSITION_MAPPING.get(vi_tri)
    
    if region_id and position_id:
        if khoi not in stats:
            stats[khoi] = {}
        if vi_tri not in stats[khoi]:
            stats[khoi][vi_tri] = 0
        stats[khoi][vi_tri] += 1

for khoi in ['Bình Dương', 'Hà Nội', 'Miền Trung', 'Hồ Chí Minh']:
    if khoi in stats:
        print(f"\n{khoi}:")
        for vi_tri, count in stats[khoi].items():
            print(f"  {vi_tri}: {count}")

print(f"\n🔑 Chạy lệnh:")
print(f"   cd /home/user/webapp && npx wrangler d1 execute webapp-production --local --file=recreate_users.sql")
